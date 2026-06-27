"""
AgriTech Startup Discovery Tracker
====================================
Uses Claude's built-in web_search tool to discover real AgriTech startups
from live news. No hardcoded data — everything is discovered at runtime.

Run:  python tracker.py
Schedule: GitHub Actions (weekly_scan.yml) runs this every Monday 08:00 UTC
"""

import json
import os
import hashlib
import time
from datetime import datetime, timezone
from typing import Optional
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()

# ── CONFIG ────────────────────────────────────────────────────────────────────

EXCEL_FILE      = "agritech_startups.xlsx"
SEEN_FILE       = "seen_hashes.json"
ANTHROPIC_API   = "https://api.anthropic.com/v1/messages"
OPENAI_API      = "https://api.openai.com/v1/chat/completions"
MODEL           = "claude-3-5-haiku-latest"
OPENAI_MODEL    = "gpt-4o-mini"

# Search queries run every scan — Claude searches the web for each one
SEARCH_QUERIES = [
    "agritech startup raised funding 2026",
    "agtech startup seed series A funding announcement 2026",
    "precision agriculture startup new product launch 2026",
    "farm robotics startup funding 2026",
    "agribiotech startup raised series 2026",
    "agri fintech startup funding announcement 2026",
    "aquaculture startup raised funding 2026",
    "livestock tech startup new product 2026",
    "food tech agri startup launch 2026",
    "smart farming startup funding round 2026",
]
ENV_FILES = [".env", "env"]

# ── ENV LOADING ──────────────────────────────────────────────────────────────

def load_env_file() -> None:
    """Load key/value pairs from a local .env or env file into the process environment."""
    for filename in ENV_FILES:
        if not os.path.exists(filename):
            continue

        with open(filename, "r", encoding="utf-8") as f:
            loaded = False
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue

                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key:
                    os.environ[key] = value
                    loaded = True

        if loaded:
            print(f"[+] Loaded environment variables from {filename}")
            return


def get_anthropic_api_keys() -> list[str]:
    """Return a list of Anthropic API keys to try, in preferred order."""
    keys = []

    raw_list = os.environ.get("ANTHROPIC_API_KEYS", "").strip()
    if raw_list:
        for value in raw_list.split(","):
            if value.strip():
                keys.append(value.strip())

    primary = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if primary and primary not in keys:
        keys.insert(0, primary)

    alt = os.environ.get("ANTHROPIC_API_KEY_ALT", "").strip()
    if alt and alt not in keys:
        keys.append(alt)

    return [key for key in keys if key]


def get_openai_api_keys() -> list[str]:
    """Return a list of OpenAI API keys to try, in preferred order."""
    keys = []

    raw_list = os.environ.get("OPENAI_API_KEYS", "").strip()
    if raw_list:
        for value in raw_list.split(","):
            if value.strip():
                keys.append(value.strip())

    primary = os.environ.get("OPENAI_API_KEY", "").strip()
    if primary and primary not in keys:
        keys.insert(0, primary)

    alt = os.environ.get("OPENAI_API_KEY_ALT", "").strip()
    if alt and alt not in keys:
        keys.append(alt)

    return [key for key in keys if key]
# ── DEDUPLICATION ─────────────────────────────────────────────────────────────

def load_seen() -> set:
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen(hashes: set):
    with open(SEEN_FILE, "w") as f:
        json.dump(sorted(hashes), f, indent=2)

def make_hash(name: str) -> str:
    return hashlib.md5(name.strip().lower().encode()).hexdigest()

# ── CLAUDE WITH WEB SEARCH ────────────────────────────────────────────────────

def call_claude_with_web_search(query: str) -> list[dict]:
    """
    Ask Claude to search the web for a query and extract AgriTech startups.
    Returns a list of startup dicts.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable not set.")

    system_prompt = """You are an expert AgriTech startup analyst.
When given a search query, use the web_search tool to find REAL, currently active
AgriTech startups that have recently launched, raised funding, or announced products.

After searching, extract each distinct startup you find and return a JSON array (no markdown fences).
Each element must have:
{
  "startup_name": "...",
  "category": "Precision Farming | Crop Tech | AgBiotech | Agri-Fintech | Farm Robotics | Supply Chain | Aquaculture | Livestock Tech | Food Tech | Other",
  "country": "...",
  "funding_stage": "Pre-Seed | Seed | Series A | Series B | Series C+ | Undisclosed | N/A",
  "description": "One sentence: what the startup does.",
  "news_summary": "One sentence: why it was recently in the news.",
  "startup_website": "https://...",
  "source_url": "https://... (the article URL you found this in)"
}

Only include startups you found evidence for in the search results.
If you find no relevant startups, return an empty array: []
Return ONLY the JSON array, nothing else."""

    payload = {
        "model": MODEL,
        "max_tokens": 1200,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"Search for: {query}"},
        ],
        "tools": [{"type": "web_search_20250305", "name": "web_search"}],
    }

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "anthropic-beta": "web-search-2025-03-05",
    }

    keys = get_anthropic_api_keys()
    if not keys:
        raise ValueError("ANTHROPIC_API_KEY is not set. Add it to the environment or env file.")

    last_error = None
    for idx, api_key in enumerate(keys, start=1):
        headers["x-api-key"] = api_key
        resp = requests.post(ANTHROPIC_API, headers=headers, json=payload, timeout=60)
        if resp.ok:
            data = resp.json()
            break

        text = resp.text
        last_error = requests.HTTPError(
            f"{resp.status_code} {resp.reason} for url: {resp.url} - {text}"
        )

        if resp.status_code == 400 and "Your credit balance is too low" in text and idx < len(keys):
            print(f"    [WARN] Anthropic key {idx} has insufficient credits, trying next key...")
            continue
        raise last_error
    else:
        raise last_error

    # Extract all text blocks from response
    text = ""
    for block in data.get("content", []):
        if block.get("type") == "text":
            text += block.get("text", "")

    # Parse the JSON array
    text = text.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    startups = json.loads(text)
    return startups if isinstance(startups, list) else []


def call_openai_with_web_search(query: str) -> list[dict]:
    """Try OpenAI as a fallback provider for startup discovery."""
    api_keys = get_openai_api_keys()
    if not api_keys:
        raise ValueError("OPENAI_API_KEY is not set. Add it to the environment or env file.")

    system_prompt = """You are an expert AgriTech startup analyst.
When given a search query, find REAL, currently active AgriTech startups that have recently launched, raised funding, or announced products.
After researching, return a JSON array (no markdown fences) with objects containing:
{
  \"startup_name\": \"...\",
  \"category\": \"Precision Farming | Crop Tech | AgBiotech | Agri-Fintech | Farm Robotics | Supply Chain | Aquaculture | Livestock Tech | Food Tech | Other\",
  \"country\": \"...\",
  \"funding_stage\": \"Pre-Seed | Seed | Series A | Series B | Series C+ | Undisclosed | N/A\",
  \"description\": \"One sentence: what the startup does.\",
  \"news_summary\": \"One sentence: why it was recently in the news.\",
  \"startup_website\": \"https://...\",
  \"source_url\": \"https://...\"
}
Return an empty array [] if no relevant startups are found."""

    payload = {
        "model": OPENAI_MODEL,
        "max_tokens": 900,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"Search for: {query}"},
        ],
        "temperature": 0,
    }

    headers = {"Content-Type": "application/json"}

    last_error = None
    for api_key in api_keys:
        headers["Authorization"] = f"Bearer {api_key}"
        resp = requests.post(OPENAI_API, headers=headers, json=payload, timeout=60)
        if resp.ok:
            data = resp.json()
            text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            text = text.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
            startups = json.loads(text)
            return startups if isinstance(startups, list) else []

        last_error = requests.HTTPError(
            f"{resp.status_code} {resp.reason} for url: {resp.url} - {resp.text}"
        )

    raise last_error if last_error else ValueError("OpenAI request failed without a response.")


def call_with_provider_fallback(query: str) -> list[dict]:
    """Try Anthropic first, then OpenAI as a fallback provider."""
    try:
        return call_claude_with_web_search(query)
    except Exception as anthropic_error:
        print(f"    [WARN] Anthropic failed: {anthropic_error}")
        try:
            print("    [WARN] Trying OpenAI fallback...")
            return call_openai_with_web_search(query)
        except Exception as openai_error:
            print(f"    [WARN] OpenAI fallback failed: {openai_error}")
            print("    [WARN] No provider available; continuing without new startup results.")
            return []


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────

COLUMNS = [
    ("Startup Name",    25),
    ("Category",        22),
    ("Country",         16),
    ("Funding Stage",   15),
    ("Description",     45),
    ("Why in News",     45),
    ("Startup Website", 35),
    ("Source URL",      40),
    ("Date Added",      18),
]

HEADER_FILL = PatternFill("solid", start_color="1F5C3A")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL    = PatternFill("solid", start_color="EAF4EE")
BODY_FONT   = Font(name="Arial", size=10)
BODY_ALIGN  = Alignment(wrap_text=True, vertical="top")
THIN        = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(bottom=THIN)

def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "AgriTech Startups"
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "AgriTech Startup Discovery Tracker"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    ws2["A3"] = "Last Updated:"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    ws2["A4"] = "Total Startups:"
    ws2["B4"] = "=COUNTA('AgriTech Startups'!A2:A10000)"
    ws2["A5"] = "Source:"
    ws2["B5"] = "Live web search via Claude AI (no hardcoded data)"

    wb.save(EXCEL_FILE)
    print(f"[+] Created new Excel file: {EXCEL_FILE}")


def append_to_excel(startups: list[dict]):
    if not startups:
        return

    wb = load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else (init_excel() or load_workbook(EXCEL_FILE))
    ws = wb["AgriTech Startups"]
    next_row = ws.max_row + 1

    for i, s in enumerate(startups):
        row = next_row + i
        values = [
            s.get("startup_name", ""),
            s.get("category", ""),
            s.get("country", "Unknown"),
            s.get("funding_stage", "N/A"),
            s.get("description", ""),
            s.get("news_summary", ""),
            s.get("startup_website", ""),
            s.get("source_url", ""),
            datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border    = CELL_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL

    wb["Summary"]["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb.save(EXCEL_FILE)
    print(f"  ✓ Saved {len(startups)} new startups → {EXCEL_FILE}")


# ── MAIN PIPELINE ─────────────────────────────────────────────────────────────

def run():
    print("=" * 60)
    print("  AgriTech Startup Discovery Tracker")
    print(f"  {datetime.now(timezone.utc).isoformat()}")
    print("=" * 60)

    load_env_file()

    if not os.path.exists(EXCEL_FILE):
        init_excel()

    seen      = load_seen()
    all_new   = []
    startup_count = 0

    for query in SEARCH_QUERIES:
        print(f"\n[→] Query: \"{query}\"")
        try:
            startups = call_with_provider_fallback(query)
            if startups:
                print(f"    Claude/OpenAI found {len(startups)} startups in search results")
                startup_count += len(startups)
            else:
                print("    [INFO] No startup results available from configured providers.")
        except Exception as e:
            print(f"    [ERROR] {e}")
            continue

        for s in startups:
            name = s.get("startup_name", "").strip()
            if not name:
                continue
            h = make_hash(name)
            if h in seen:
                print(f"    [SKIP] Already tracked: {name}")
                continue
            seen.add(h)
            all_new.append(s)
            print(f"    [NEW]  {name} ({s.get('country','?')}) — {s.get('funding_stage','?')}")

        time.sleep(1)  # avoid rate limits

    save_seen(seen)

    if all_new:
        print(f"\n[+] Discovered {len(all_new)} new startups this scan.")
        append_to_excel(all_new)
    else:
        print("\n[=] No new startups found this scan.")

    print(f"\n[✓] Done. Excel file: {EXCEL_FILE}")


if __name__ == "__main__":
    run()
